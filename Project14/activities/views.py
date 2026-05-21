from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.db.models import Count
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from .models import Activity, Registration, CheckIn
from .forms import ActivityForm, RegistrationForm
import datetime


def home(request):
    now = timezone.now()
    upcoming_activities = Activity.objects.filter(status='upcoming')[:5]
    ongoing_activities = Activity.objects.filter(status='ongoing')[:5]
    context = {
        'upcoming_activities': upcoming_activities,
        'ongoing_activities': ongoing_activities,
    }
    return render(request, 'home.html', context)


def activity_list(request):
    activities = Activity.objects.all()
    status_filter = request.GET.get('status', '')
    if status_filter:
        activities = activities.filter(status=status_filter)
    context = {
        'activities': activities,
        'status_filter': status_filter,
    }
    return render(request, 'activities/list.html', context)


def activity_detail(request, pk):
    activity = get_object_or_404(Activity, pk=pk)
    is_registered = False
    if request.method == 'POST':
        form = RegistrationForm(request.POST)
        if form.is_valid():
            if not activity.can_register():
                messages.error(request, '该活动报名已截止或名额已满')
                return redirect('activity_detail', pk=pk)
            
            student_id = form.cleaned_data['student_id']
            if Registration.objects.filter(activity=activity, student_id=student_id).exists():
                messages.error(request, '该学号已报名此活动')
                return redirect('activity_detail', pk=pk)
            
            registration = form.save(commit=False)
            registration.activity = activity
            registration.save()
            activity.current_participants += 1
            activity.save()
            messages.success(request, '报名成功！')
            return redirect('activity_detail', pk=pk)
    else:
        form = RegistrationForm()
    
    registrations = Registration.objects.filter(activity=activity)
    checkin_count = CheckIn.objects.filter(registration__activity=activity).count()
    
    context = {
        'activity': activity,
        'form': form,
        'registrations': registrations,
        'checkin_count': checkin_count,
    }
    return render(request, 'activities/detail.html', context)


@login_required
def activity_create(request):
    if request.method == 'POST':
        form = ActivityForm(request.POST, request.FILES)
        if form.is_valid():
            activity = form.save(commit=False)
            activity.created_by = request.user
            activity.save()
            messages.success(request, '活动发布成功！')
            return redirect('activity_detail', pk=activity.pk)
    else:
        form = ActivityForm()
    return render(request, 'activities/form.html', {'form': form, 'title': '发布活动'})


@login_required
def activity_edit(request, pk):
    activity = get_object_or_404(Activity, pk=pk)
    if request.method == 'POST':
        form = ActivityForm(request.POST, request.FILES, instance=activity)
        if form.is_valid():
            form.save()
            messages.success(request, '活动更新成功！')
            return redirect('activity_detail', pk=pk)
    else:
        form = ActivityForm(instance=activity)
    return render(request, 'activities/form.html', {'form': form, 'title': '编辑活动'})


@login_required
def activity_delete(request, pk):
    activity = get_object_or_404(Activity, pk=pk)
    activity.delete()
    messages.success(request, '活动已删除')
    return redirect('activity_list')


@login_required
def registration_cancel(request, pk):
    registration = get_object_or_404(Registration, pk=pk)
    activity_pk = registration.activity.pk
    activity = registration.activity
    activity.current_participants -= 1
    activity.save()
    registration.delete()
    messages.success(request, '报名已取消')
    return redirect('activity_detail', pk=activity_pk)


@login_required
def checkin_view(request, activity_pk):
    activity = get_object_or_404(Activity, pk=activity_pk)
    registrations = Registration.objects.filter(activity=activity)
    
    if request.method == 'POST':
        student_id = request.POST.get('student_id', '').strip()
        try:
            registration = Registration.objects.get(activity=activity, student_id=student_id)
            if hasattr(registration, 'checkin'):
                messages.warning(request, '该学生已签到')
            else:
                CheckIn.objects.create(registration=registration, checked_in_by=request.user)
                messages.success(request, f'{registration.student_name} 签到成功')
        except Registration.DoesNotExist:
            messages.error(request, '未找到该学生的报名记录')
    
    checked_in_ids = CheckIn.objects.filter(registration__activity=activity).values_list('registration_id', flat=True)
    
    context = {
        'activity': activity,
        'registrations': registrations,
        'checked_in_ids': checked_in_ids,
    }
    return render(request, 'activities/checkin.html', context)


@login_required
def export_registrations(request, pk):
    activity = get_object_or_404(Activity, pk=pk)
    registrations = Registration.objects.filter(activity=activity).select_related('checkin')
    
    wb = Workbook()
    ws = wb.active
    ws.title = '报名名单'
    
    headers = ['序号', '学生姓名', '学号', '学院', '联系电话', '邮箱', '报名时间', '签到状态', '签到时间', '备注']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    for row_num, reg in enumerate(registrations, 2):
        has_checkin = hasattr(reg, 'checkin')
        ws.cell(row=row_num, column=1, value=row_num - 1)
        ws.cell(row=row_num, column=2, value=reg.student_name)
        ws.cell(row=row_num, column=3, value=reg.student_id)
        ws.cell(row=row_num, column=4, value=reg.college)
        ws.cell(row=row_num, column=5, value=reg.phone)
        ws.cell(row=row_num, column=6, value=reg.email or '')
        ws.cell(row=row_num, column=7, value=reg.registered_at.strftime('%Y-%m-%d %H:%M:%S'))
        ws.cell(row=row_num, column=8, value='已签到' if has_checkin else '未签到')
        ws.cell(row=row_num, column=9, value=reg.checkin.checked_in_at.strftime('%Y-%m-%d %H:%M:%S') if has_checkin else '')
        ws.cell(row=row_num, column=10, value=reg.remarks or '')
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{activity.title}_报名名单.xlsx"'
    wb.save(response)
    return response


@login_required
def dashboard(request):
    now = timezone.now()
    total_activities = Activity.objects.count()
    total_registrations = Registration.objects.count()
    total_checkins = CheckIn.objects.count()
    upcoming_count = Activity.objects.filter(status='upcoming').count()
    ongoing_count = Activity.objects.filter(status='ongoing').count()
    ended_count = Activity.objects.filter(status='ended').count()
    
    recent_activities = Activity.objects.all()[:10]
    
    activity_stats = Registration.objects.values('activity__title').annotate(
        count=Count('id')
    ).order_by('-count')[:5]
    
    context = {
        'total_activities': total_activities,
        'total_registrations': total_registrations,
        'total_checkins': total_checkins,
        'upcoming_count': upcoming_count,
        'ongoing_count': ongoing_count,
        'ended_count': ended_count,
        'recent_activities': recent_activities,
        'activity_stats': activity_stats,
    }
    return render(request, 'dashboard.html', context)
